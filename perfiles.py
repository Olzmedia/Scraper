"""
Gestión de perfiles. Cada perfil tiene su propia carpeta con:
  · config.json  (la configuración guardada del perfil)
  · temp_csv/    (los leads extraídos, un CSV por nicho)
  · leads.xlsx   (exportación)
Todos los perfiles viven bajo <datos>/perfiles/<nombre>/.
"""
import os
import re
import csv
import json
import shutil

import scraper_google_maps as motor


def _base():
    d = os.path.join(motor.dir_datos(), "perfiles")
    os.makedirs(d, exist_ok=True)
    return d


def _slug(nombre):
    s = re.sub(r"[^\w\- ]", "", nombre).strip().replace(" ", "_")
    return s[:40] or "perfil"


def ruta_perfil(nombre):
    return os.path.join(_base(), _slug(nombre))


def listar():
    """Lista de perfiles existentes (por su nombre mostrado)."""
    base = _base()
    salida = []
    for d in sorted(os.listdir(base)):
        meta = os.path.join(base, d, "perfil.json")
        if os.path.isfile(meta):
            try:
                with open(meta, encoding="utf-8") as f:
                    salida.append(json.load(f).get("nombre", d))
            except Exception:
                salida.append(d)
    return salida


def crear(nombre):
    nombre = (nombre or "").strip()
    if not nombre:
        raise ValueError("El nombre del perfil no puede estar vacío.")
    p = ruta_perfil(nombre)
    if os.path.isdir(p):
        raise ValueError("Ya existe un perfil con ese nombre.")
    os.makedirs(os.path.join(p, "temp_csv"), exist_ok=True)
    with open(os.path.join(p, "perfil.json"), "w", encoding="utf-8") as f:
        json.dump({"nombre": nombre}, f, ensure_ascii=False)
    guardar_config(nombre, {})
    return nombre


def renombrar(viejo, nuevo):
    nuevo = (nuevo or "").strip()
    if not nuevo:
        raise ValueError("El nuevo nombre no puede estar vacío.")
    pv = ruta_perfil(viejo)
    pn = ruta_perfil(nuevo)
    if not os.path.isdir(pv):
        raise ValueError("El perfil a renombrar no existe.")
    if os.path.isdir(pn) and _slug(viejo) != _slug(nuevo):
        raise ValueError("Ya existe un perfil con ese nombre.")
    if _slug(viejo) != _slug(nuevo):
        os.rename(pv, pn)
    with open(os.path.join(pn, "perfil.json"), "w", encoding="utf-8") as f:
        json.dump({"nombre": nuevo}, f, ensure_ascii=False)
    return nuevo


def borrar_perfil(nombre):
    p = ruta_perfil(nombre)
    if os.path.isdir(p):
        shutil.rmtree(p, ignore_errors=True)
    return True


def borrar_leads(nombre):
    """Borra solo los leads (temp_csv y excel), conserva la config."""
    p = ruta_perfil(nombre)
    shutil.rmtree(os.path.join(p, "temp_csv"), ignore_errors=True)
    os.makedirs(os.path.join(p, "temp_csv"), exist_ok=True)
    x = os.path.join(p, "leads.xlsx")
    if os.path.isfile(x):
        try:
            os.remove(x)
        except Exception:
            pass
    return True


def guardar_config(nombre, cfg):
    p = ruta_perfil(nombre)
    os.makedirs(p, exist_ok=True)
    with open(os.path.join(p, "config.json"), "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


def cargar_config(nombre):
    p = os.path.join(ruta_perfil(nombre), "config.json")
    if os.path.isfile(p):
        try:
            with open(p, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def carpeta_temp(nombre):
    return os.path.join(ruta_perfil(nombre), "temp_csv")


def ruta_excel(nombre):
    return os.path.join(ruta_perfil(nombre), "leads.xlsx")


def cargar_leads(nombre):
    """Lee todos los leads guardados del perfil desde sus CSV."""
    tmp = carpeta_temp(nombre)
    filas = []
    if not os.path.isdir(tmp):
        return filas
    for archivo in sorted(os.listdir(tmp)):
        if not archivo.endswith(".csv"):
            continue
        nicho = archivo[:-4]
        try:
            with open(os.path.join(tmp, archivo), encoding="utf-8-sig") as f:
                for fila in csv.DictReader(f):
                    fila["_nicho"] = nicho
                    filas.append(fila)
        except Exception:
            continue
    return filas


def estadisticas(nombre):
    """Resumen y datos para las gráficas del panel."""
    leads = cargar_leads(nombre)
    total = len(leads)
    por_nicho, por_zona = {}, {}
    score_buckets = {"Alto (≥5)": 0, "Medio (3-5)": 0, "Bajo (<3)": 0}
    cob = {"whatsapp": 0, "email_valido": 0, "web": 0, "sin_web": 0, "telefono": 0}
    for l in leads:
        n = l.get("_nicho") or l.get("categoria") or "otros"
        por_nicho[n] = por_nicho.get(n, 0) + 1
        z = l.get("zona") or "—"
        por_zona[z] = por_zona.get(z, 0) + 1
        try:
            sc = float(l.get("score") or 0)
        except Exception:
            sc = 0
        if sc >= 5:
            score_buckets["Alto (≥5)"] += 1
        elif sc >= 3:
            score_buckets["Medio (3-5)"] += 1
        else:
            score_buckets["Bajo (<3)"] += 1
        if l.get("telefono"):
            cob["telefono"] += 1
        if l.get("whatsapp"):
            cob["whatsapp"] += 1
        if l.get("email_estado") == "válido":
            cob["email_valido"] += 1
        if l.get("sin_web"):
            cob["sin_web"] += 1
        else:
            cob["web"] += 1
    return {
        "total": total,
        "por_nicho": por_nicho,
        "por_zona": por_zona,
        "score": score_buckets,
        "cobertura": cob,
    }


# ---------- tema (colores) por perfil ----------
def guardar_tema(nombre, tema):
    """tema = {acento:int, fondo:int, terminal:bool}"""
    p = ruta_perfil(nombre)
    os.makedirs(p, exist_ok=True)
    with open(os.path.join(p, "tema.json"), "w", encoding="utf-8") as f:
        json.dump(tema or {}, f, ensure_ascii=False)


def cargar_tema(nombre):
    p = os.path.join(ruta_perfil(nombre), "tema.json")
    if os.path.isfile(p):
        try:
            with open(p, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


# ---------- foto de perfil (guardada como base64 dataURL) ----------
def guardar_foto(nombre, data_url):
    p = ruta_perfil(nombre)
    os.makedirs(p, exist_ok=True)
    with open(os.path.join(p, "foto.txt"), "w", encoding="utf-8") as f:
        f.write(data_url or "")
    return True


def cargar_foto(nombre):
    p = os.path.join(ruta_perfil(nombre), "foto.txt")
    if os.path.isfile(p):
        try:
            with open(p, encoding="utf-8") as f:
                return f.read()
        except Exception:
            return ""
    return ""


def quitar_foto(nombre):
    p = os.path.join(ruta_perfil(nombre), "foto.txt")
    if os.path.isfile(p):
        try:
            os.remove(p)
        except Exception:
            pass
    return True


# ---------- ediciones de la hoja (overrides + columnas propias + eliminados) ----------
def cargar_hoja(nombre):
    """Devuelve {ediciones:{url:{campo:valor}}, columnas:[...], eliminados:[url,...]}."""
    p = os.path.join(ruta_perfil(nombre), "hoja.json")
    if os.path.isfile(p):
        try:
            with open(p, encoding="utf-8") as f:
                d = json.load(f)
                d.setdefault("ediciones", {})
                d.setdefault("columnas", [])
                d.setdefault("eliminados", [])
                return d
        except Exception:
            pass
    return {"ediciones": {}, "columnas": [], "eliminados": []}


def guardar_hoja(nombre, hoja):
    p = ruta_perfil(nombre)
    os.makedirs(p, exist_ok=True)
    with open(os.path.join(p, "hoja.json"), "w", encoding="utf-8") as f:
        json.dump(hoja or {}, f, ensure_ascii=False, indent=2)
    return True


def leads_editados(nombre):
    """Aplica ediciones, columnas propias y eliminados sobre los leads del perfil."""
    base = cargar_leads(nombre)
    h = cargar_hoja(nombre)
    ed, cols, elim = h["ediciones"], h["columnas"], set(h["eliminados"])
    salida = []
    for l in base:
        u = l.get("url_maps", "")
        if u in elim:
            continue
        fila = dict(l)
        for c in cols:
            fila.setdefault(c, "")
        if u in ed:
            fila.update(ed[u])
        salida.append(fila)
    return salida, cols


def exportar_hoja(nombre, destino_dir, visibles):
    """Construye un Excel con los leads editados (columnas visibles + propias)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    leads, extra = leads_editados(nombre)
    columnas = list(visibles) + [c for c in extra if c not in visibles]
    headers_map = dict(zip(motor.CAMPOS, motor.ENCABEZADOS))

    os.makedirs(destino_dir, exist_ok=True)
    archivo = os.path.join(destino_dir, f"leads_{_slug(nombre)}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    fill = PatternFill("solid", fgColor="6D5AE6")
    fnt = Font(bold=True, color="FFFFFF")

    encabezados = [headers_map.get(c, c.capitalize()) for c in columnas]
    ws.append(encabezados)
    for i in range(1, len(encabezados) + 1):
        c = ws.cell(row=1, column=i)
        c.fill = fill; c.font = fnt; c.alignment = Alignment(vertical="center")
    for l in leads:
        ws.append([l.get(c, "") for c in columnas])

    anchos = {"nombre": 28, "direccion": 34, "sitio_web": 30, "email": 26,
              "url_maps": 40, "notas": 30, "estado": 16}
    for i, c in enumerate(columnas, 1):
        ws.column_dimensions[get_column_letter(i)].width = anchos.get(c, 16)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columnas))}{ws.max_row}"
    wb.save(archivo)
    return archivo


# ---------- leads marcados como contactados ----------
def cargar_contactados(nombre):
    p = os.path.join(ruta_perfil(nombre), "contactados.json")
    if os.path.isfile(p):
        try:
            with open(p, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []


def guardar_contactados(nombre, lista):
    p = ruta_perfil(nombre)
    os.makedirs(p, exist_ok=True)
    with open(os.path.join(p, "contactados.json"), "w", encoding="utf-8") as f:
        json.dump(lista or [], f, ensure_ascii=False)
    return True


# ---------- plantillas de mensaje (WhatsApp / correo) por perfil ----------
_MSJ_DEF = {
    "wa": "Hola {nombre}, ¿cómo están? Somos una agencia local y ayudamos a "
          "negocios como el suyo a conseguir más clientes por internet. "
          "¿Les interesaría que les cuente cómo, sin compromiso?",
    "email_asunto": "Más clientes para {nombre}",
    "email_cuerpo": "Hola {nombre},\n\n"
                    "Le escribo de [TU NEGOCIO]. Ayudamos a negocios de la zona a "
                    "conseguir más clientes por internet (redes, Google, publicidad).\n\n"
                    "Me gustaría mostrarle en pocos minutos, sin compromiso, cómo podríamos "
                    "ayudarle a {nombre}. ¿Tendría un momento esta semana?\n\n"
                    "Quedo atento. Saludos,\n[TU NOMBRE]\n[TU TELÉFONO / WHATSAPP]",
    "email_proveedor": "gmail",
}


def cargar_mensajes(nombre):
    p = os.path.join(ruta_perfil(nombre), "mensajes.json")
    if os.path.isfile(p):
        try:
            with open(p, encoding="utf-8") as f:
                d = json.load(f)
                for k, v in _MSJ_DEF.items():
                    d.setdefault(k, v)
                return d
        except Exception:
            pass
    return dict(_MSJ_DEF)


def guardar_mensajes(nombre, msjs):
    p = ruta_perfil(nombre)
    os.makedirs(p, exist_ok=True)
    with open(os.path.join(p, "mensajes.json"), "w", encoding="utf-8") as f:
        json.dump(msjs or {}, f, ensure_ascii=False, indent=2)
    return True


# ---------- estado de contacto por canal (wa / email) ----------
def cargar_enviados(nombre):
    """Devuelve {wa:[url,...], email:[url,...]} de a quién ya se escribió."""
    p = os.path.join(ruta_perfil(nombre), "enviados.json")
    if os.path.isfile(p):
        try:
            with open(p, encoding="utf-8") as f:
                d = json.load(f)
                d.setdefault("wa", [])
                d.setdefault("email", [])
                return d
        except Exception:
            pass
    return {"wa": [], "email": []}


def guardar_enviados(nombre, datos):
    p = ruta_perfil(nombre)
    os.makedirs(p, exist_ok=True)
    with open(os.path.join(p, "enviados.json"), "w", encoding="utf-8") as f:
        json.dump(datos or {"wa": [], "email": []}, f, ensure_ascii=False)
    return True
