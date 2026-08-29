"""
============================================================
 LEAD SCRAPER  ·  Google Maps  ·  motor
============================================================
Extrae leads de negocios por múltiples nichos y zonas, en paralelo,
y los entrega en un Excel profesional priorizado.

Mejoras de captación:
  · Sinónimos por nicho (varios términos de búsqueda -> más resultados)
  · WhatsApp, Instagram y Facebook además de email y teléfono
  · Cuadrícula automática de zonas (cobertura sistemática por radio)
  · Puntaje de lead y ordenamiento del mejor al peor
  · Normalización de teléfonos (formato +57) y marca de "sin web"

Robustez:
  · Selectores con respaldo · reintentos · reanudar (resume)
  · Detección de captcha con pausa/enfriamiento · escritura incremental

NO usa la API de Google Places.
"""

import asyncio
import csv
import os
import re
import sys
import math
import time
import random
import logging
import threading
import subprocess

import yaml
import httpx
import verificacion
from playwright.async_api import async_playwright
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------- columnas (orden pensado para prospección) ----------
CAMPOS = ["nombre", "categoria", "score", "telefono", "whatsapp",
          "whatsapp_estado", "email", "email_estado", "instagram", "facebook",
          "sitio_web", "web_activa", "sin_web", "actividad", "calificacion",
          "resenas", "direccion", "horario", "coordenadas", "zona", "url_maps"]
ENCABEZADOS = ["Nombre", "Categoría", "Score", "Teléfono", "WhatsApp",
               "WA estado", "Email", "Email estado", "Instagram", "Facebook",
               "Sitio web", "Web activa", "Sin web", "Actividad", "Calificación",
               "Reseñas", "Dirección", "Horario", "Coordenadas", "Zona", "URL Maps"]

USER_AGENT = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
              "AppleWebKit/537.36 (KHTML, like Gecko) "
              "Chrome/124.0.0.0 Safari/537.36")

# Parches para que el navegador parezca más humano (reduce captchas)
STEALTH_JS = """
Object.defineProperty(navigator,'webdriver',{get:()=>undefined});
window.chrome={runtime:{}};
Object.defineProperty(navigator,'languages',{get:()=>['es-ES','es']});
Object.defineProperty(navigator,'plugins',{get:()=>[1,2,3,4,5]});
const _q=window.navigator.permissions&&window.navigator.permissions.query;
if(_q){window.navigator.permissions.query=(p)=>p&&p.name==='notifications'
  ?Promise.resolve({state:Notification.permission}):_q(p);}
"""

# Modo prudente (más lento, menos bloqueos)
LENTO = False

RE_EMAIL = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
RE_WA = re.compile(r"(?:wa\.me/|api\.whatsapp\.com/send\?phone=|"
                   r"whatsapp\.com/send\?phone=)(\+?\d[\d]{6,15})")
RE_IG = re.compile(r"instagram\.com/([A-Za-z0-9_.]+)")
RE_FB = re.compile(r"facebook\.com/([A-Za-z0-9_.\-]+)")
EMAIL_BASURA = ("sentry", "wixpress", "example.com", "domain.com", ".png",
                ".jpg", ".jpeg", ".gif", ".webp", "@2x")
IG_BASURA = {"explore", "p", "reel", "reels", "stories", "accounts", "about",
             "developer", "legal", "directory"}
FB_BASURA = {"sharer", "sharer.php", "plugins", "tr", "dialog", "profile.php",
             "pages", "people", "groups", "watch", "events"}

log = logging.getLogger("scraper")

# ---------- evento de parada (lo usa la interfaz gráfica) ----------
PARAR = threading.Event()

# ---------- gancho: se llama con cada lead nuevo (lo usa el dashboard) ----------
ON_LEAD = None
PROGRESO = {"fase": "", "total": 0}


# ---------- carpeta de datos y navegador (para el .exe) ----------
def dir_datos():
    base = os.environ.get("LOCALAPPDATA") or os.path.expanduser("~")
    d = os.path.join(base, "LeadScraper")
    os.makedirs(d, exist_ok=True)
    return d


BROWSERS_DIR = os.path.join(dir_datos(), "browsers")
os.environ.setdefault("PLAYWRIGHT_BROWSERS_PATH", BROWSERS_DIR)


def navegador_instalado():
    if not os.path.isdir(BROWSERS_DIR):
        return False
    return any(n.startswith("chromium") for n in os.listdir(BROWSERS_DIR))


def instalar_navegador(log_cb=None):
    def _log(msg):
        (log_cb or log.info)(msg)
    os.makedirs(BROWSERS_DIR, exist_ok=True)
    os.environ["PLAYWRIGHT_BROWSERS_PATH"] = BROWSERS_DIR
    from playwright._impl._driver import compute_driver_executable, get_driver_env
    node, cli = compute_driver_executable()
    _log("Descargando el navegador Chromium (~150 MB, solo la primera vez)...")
    proc = subprocess.Popen([node, cli, "install", "chromium"],
                            env=get_driver_env(),
                            stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                            text=True)
    for linea in proc.stdout:
        _log("  " + linea.rstrip())
    proc.wait()
    ok = proc.returncode == 0
    _log("✓ Navegador instalado." if ok else "✗ Falló la instalación.")
    return ok


# ============================================================
#  CONFIGURACIÓN
# ============================================================
def cargar_config(ruta):
    if not os.path.exists(ruta):
        log.error(f"No se encontró la configuración: {ruta}")
        sys.exit(1)
    with open(ruta, encoding="utf-8") as f:
        return yaml.safe_load(f)


def configurar_logging():
    logging.basicConfig(
        level=logging.INFO, format="%(asctime)s  %(message)s",
        datefmt="%H:%M:%S",
        handlers=[logging.StreamHandler(),
                  logging.FileHandler("scraper.log", encoding="utf-8")])


def parse_nichos(lista):
    """Cada entrada puede traer sinónimos separados por '|'.
    Devuelve [(canonico, [terminos...]), ...]."""
    salida = []
    for entrada in lista:
        partes = [p.strip() for p in str(entrada).split("|") if p.strip()]
        if not partes:
            continue
        salida.append((partes[0], partes))
    return salida


def generar_grilla(lat, lng, radio_km, filas, columnas):
    """Genera una malla de coordenadas alrededor de un centro."""
    zonas = []
    dlat = radio_km / 111.0
    coslat = math.cos(math.radians(lat)) or 1e-6
    dlng = radio_km / (111.0 * coslat)
    for i in range(filas):
        fy = 0 if filas == 1 else (i - (filas - 1) / 2) / ((filas - 1) / 2)
        for j in range(columnas):
            fx = 0 if columnas == 1 else (j - (columnas - 1) / 2) / ((columnas - 1) / 2)
            zonas.append({"nombre": f"celda_{i+1}_{j+1}",
                          "lat": round(lat + fy * dlat, 6),
                          "lng": round(lng + fx * dlng, 6)})
    return zonas


def resolver_zonas(b):
    grilla = b.get("grilla") or {}
    if grilla.get("usar"):
        return generar_grilla(grilla["centro_lat"], grilla["centro_lng"],
                              grilla.get("radio_km", 6),
                              grilla.get("filas", 3),
                              grilla.get("columnas", 3))
    return b["zonas"]


# ============================================================
#  UTILIDADES
# ============================================================
async def _texto(page, selectores):
    for sel in selectores:
        try:
            el = await page.query_selector(sel)
            if el:
                t = (await el.inner_text()).strip()
                if t:
                    return t
        except Exception:
            continue
    return ""


async def _attr(page, selectores, atributo):
    for sel in selectores:
        try:
            el = await page.query_selector(sel)
            if el:
                v = (await el.get_attribute(atributo) or "").strip()
                if v:
                    return v
        except Exception:
            continue
    return ""


def limpiar_aria(valor, prefijos):
    for p in prefijos:
        if valor.lower().startswith(p.lower()):
            return valor[len(p):].strip(": ").strip()
    return valor.strip()


def clave_lugar(href):
    m = re.search(r"/place/([^/]+)", href or "")
    return m.group(1) if m else href


def nombre_hoja(texto):
    limpio = re.sub(r'[\\/*?:\[\]]', "", texto)
    return limpio[:31] if limpio else "Hoja"


def normalizar_tel(t):
    """Normaliza a formato colombiano +57 cuando aplica."""
    if not t:
        return ""
    d = re.sub(r"[^\d]", "", t)
    if d.startswith("57") and len(d) >= 12:
        d = d[2:]
    if len(d) == 10 and d.startswith("3"):
        return f"+57 {d[:3]} {d[3:]}"
    if len(d) == 7:  # fijo local
        return f"{d[:3]} {d[3:]}"
    return t.strip()


def calcular_score(f):
    s = 0.0
    if f.get("whatsapp"): s += 3
    if f.get("telefono"): s += 2
    if f.get("email"): s += 2
    if f.get("instagram") or f.get("facebook"): s += 1
    # bonus por contactos VERIFICADOS
    if f.get("email_estado") == "válido": s += 1
    if f.get("whatsapp_estado") == "válido": s += 1
    if f.get("actividad") == "activo": s += 1
    try:
        rev = int(re.sub(r"[^\d]", "", f.get("resenas") or "0") or 0)
    except Exception:
        rev = 0
    s += min(rev, 200) / 100.0            # hasta +2 por reputación
    try:
        cal = float(f.get("calificacion") or 0)
    except Exception:
        cal = 0
    if cal:
        s += max(0, cal - 3)              # nudge por buena calificación
    return round(s, 1)


async def aceptar_consentimiento(page):
    """Si aparece la pantalla o el diálogo de consentimiento de Google, lo acepta.
    Es una de las causas más comunes de 'no llegan resultados': con cookies nuevas,
    Google redirige a consent.google.com y sin aceptar no carga el mapa."""
    selectores = [
        'button[aria-label="Aceptar todo"]',
        'button[aria-label="Rechazar todo"]',
        'button[aria-label="Accept all"]',
        'button[aria-label="Reject all"]',
        'form[action*="consent"] button[jsname]',
        'button:has-text("Aceptar todo")',
        'button:has-text("Rechazar todo")',
        'button:has-text("Acepto")',
        'button:has-text("Accept all")',
    ]
    hizo_click = False
    for sel in selectores:
        try:
            btn = await page.query_selector(sel)
            if btn:
                await btn.click()
                hizo_click = True
                await page.wait_for_timeout(1500)
                break
        except Exception:
            continue
    if hizo_click:
        try:
            await page.wait_for_load_state("domcontentloaded", timeout=8000)
        except Exception:
            pass
    return hizo_click


async def detectar_bloqueo(page):
    if "/sorry/" in page.url or "consent.google" in page.url:
        return True
    try:
        if await page.query_selector('form#captcha-form, iframe[src*="recaptcha"]'):
            return True
    except Exception:
        pass
    return False


async def manejar_bloqueo(page, cfg):
    headless = cfg["rendimiento"].get("headless", False)
    enfriamiento = cfg["rendimiento"].get("enfriamiento_seg", 180)
    log.warning("=" * 55)
    log.warning("⚠️  Google mostró un CAPTCHA / bloqueo.")
    if not headless:
        log.warning("   Resuélvelo en la ventana. El script espera 3 min...")
        for _ in range(90):
            await page.wait_for_timeout(2000)
            if not await detectar_bloqueo(page):
                log.info("   ✓ Resuelto, continuando.")
                return True
        return False
    # headless: enfriamiento automático
    log.warning(f"   Modo oculto: enfriando {enfriamiento}s antes de reintentar.")
    await page.wait_for_timeout(enfriamiento * 1000)
    return not await detectar_bloqueo(page)


async def con_reintentos(coro_func, reintentos, desc=""):
    for intento in range(1, reintentos + 1):
        try:
            return await coro_func()
        except Exception as e:
            if intento == reintentos:
                log.warning(f"  Falló ({desc}): {e}")
                return None
            await asyncio.sleep(2 * intento + random.random())
    return None


# ============================================================
#  CONTACTOS desde la web del negocio (email, WhatsApp, redes)
# ============================================================
def _limpiar_ig(user):
    user = user.strip("/").split("/")[0].split("?")[0]
    return "" if user.lower() in IG_BASURA or not user else user


def _limpiar_fb(user):
    user = user.strip("/").split("?")[0]
    return "" if user.lower() in FB_BASURA or not user else user


async def extraer_contactos(cliente, url_web, timeout, quiere_redes):
    """Devuelve dict con email/whatsapp/instagram/facebook desde la web."""
    res = {"email": "", "whatsapp": "", "instagram": "", "facebook": ""}
    if not url_web:
        return res
    candidatas = [url_web]
    for extra in ("/contacto", "/contact", "/contactenos"):
        candidatas.append(url_web.rstrip("/") + extra)

    emails, wa, ig, fb = [], [], [], []
    for u in candidatas:
        try:
            r = await cliente.get(u, timeout=timeout, follow_redirects=True)
            html = r.text
        except Exception:
            continue
        for m in RE_EMAIL.findall(html):
            low = m.lower()
            if not any(x in low for x in EMAIL_BASURA) and m not in emails:
                emails.append(m)
        if quiere_redes:
            for m in RE_WA.findall(html):
                if m not in wa:
                    wa.append(m)
            for m in RE_IG.findall(html):
                c = _limpiar_ig(m)
                if c and c not in ig:
                    ig.append(c)
            for m in RE_FB.findall(html):
                c = _limpiar_fb(m)
                if c and c not in fb:
                    fb.append(c)
        if emails and (not quiere_redes or wa or ig or fb):
            break

    res["email"] = "; ".join(emails[:2])
    if wa:
        res["whatsapp"] = normalizar_tel(wa[0])
    if ig:
        res["instagram"] = "instagram.com/" + ig[0]
    if fb:
        res["facebook"] = "facebook.com/" + fb[0]
    return res


# ============================================================
#  ESCRITOR INCREMENTAL + CHECKPOINT
# ============================================================
class EscritorCSV:
    def __init__(self, carpeta):
        os.makedirs(carpeta, exist_ok=True)
        self.carpeta = carpeta
        self.handles = {}
        self.conteo = {}
        self.total = 0
        self.ya_guardados = set()
        self._cargar_existentes()

    def _ruta(self, nicho):
        return os.path.join(self.carpeta, nombre_hoja(nicho) + ".csv")

    def _cargar_existentes(self):
        for archivo in os.listdir(self.carpeta) if os.path.isdir(self.carpeta) else []:
            if not archivo.endswith(".csv"):
                continue
            try:
                with open(os.path.join(self.carpeta, archivo),
                          encoding="utf-8-sig") as f:
                    for fila in csv.DictReader(f):
                        self.ya_guardados.add(clave_lugar(fila.get("url_maps", "")))
            except Exception:
                continue
        if self.ya_guardados:
            log.info(f"↻ Reanudando: {len(self.ya_guardados)} ya guardados, se omiten.")

    def _writer(self, nicho):
        if nicho not in self.handles:
            ruta = self._ruta(nicho)
            existe = os.path.exists(ruta)
            f = open(ruta, "a", newline="", encoding="utf-8-sig")
            w = csv.DictWriter(f, fieldnames=CAMPOS)
            if not existe:
                w.writeheader()
            self.handles[nicho] = (f, w)
            self.conteo[nicho] = 0
        return self.handles[nicho][1]

    def ya_existe(self, href):
        return clave_lugar(href) in self.ya_guardados

    def escribir(self, nicho, fila):
        self._writer(nicho).writerow({k: fila.get(k, "") for k in CAMPOS})
        self.handles[nicho][0].flush()
        self.conteo[nicho] = self.conteo.get(nicho, 0) + 1
        self.total += 1
        self.ya_guardados.add(clave_lugar(fila.get("url_maps", "")))

    def cerrar(self):
        for f, _ in self.handles.values():
            f.close()


# ============================================================
#  FASE 1 — recolectar enlaces (nichos x sinónimos x zonas)
# ============================================================
async def recolectar_en_zona(page, termino, zona, faltan, zoom):
    query = termino.replace(" ", "+")
    url = (f"https://www.google.com/maps/search/{query}/"
           f"@{zona['lat']},{zona['lng']},{zoom}?hl=es")
    await page.goto(url, wait_until="domcontentloaded")
    await aceptar_consentimiento(page)
    if await detectar_bloqueo(page):
        return "BLOQUEO"
    await page.wait_for_timeout(2000 + random.randint(0, 600))
    enlaces = []
    try:
        await page.wait_for_selector('div[role="feed"]', timeout=15000)
    except Exception:
        # Caso 1: Google abrió directamente una sola ficha (resultado único).
        if "/maps/place/" in page.url:
            log.info(f"  Resultado único directo para '{termino}' @ {zona['nombre']}")
            return [page.url]
        # Caso 2: no cargó el feed (layout distinto, sin resultados,
        # consentimiento o bloqueo). Lo anotamos y guardamos captura.
        log.warning(f"  ⚠ No cargó la lista para '{termino}' @ {zona['nombre']} "
                    f"(url actual: {page.url})")
        try:
            dbg = os.path.join(dir_datos(), "debug")
            os.makedirs(dbg, exist_ok=True)
            await page.screenshot(
                path=os.path.join(dbg, f"sin_lista_{int(time.time())}.png"))
        except Exception:
            pass
        return enlaces
    feed = await page.query_selector('div[role="feed"]')
    intentos = 0
    while len(enlaces) < faltan and intentos < 6:
        tarjetas = await page.query_selector_all('a.hfpxzc')
        if not tarjetas:  # respaldo por si Google cambió la clase de la tarjeta
            tarjetas = await page.query_selector_all(
                'div[role="feed"] a[href*="/maps/place/"]')
        for c in tarjetas:
            href = await c.get_attribute("href")
            if href and href not in enlaces:
                enlaces.append(href)
        if await page.query_selector('span:has-text("Llegaste al final de la lista")'):
            break
        antes = len(enlaces)
        if feed:
            await page.evaluate("(el)=>el.scrollBy(0,1000+Math.random()*600)", feed)
        await page.wait_for_timeout(1400 + random.randint(0, 900))
        intentos = intentos + 1 if len(enlaces) == antes else 0
    return enlaces


async def worker_fase1(page, cola, acumulado, lock, cfg):
    target = cfg["busqueda"]["target_por_nicho"]
    zoom = cfg["busqueda"].get("zoom", "13z")
    reintentos = cfg["rendimiento"].get("reintentos", 3)
    while True:
        if PARAR.is_set():
            return
        try:
            canonico, termino, zona = cola.get_nowait()
        except asyncio.QueueEmpty:
            return
        async with lock:
            ya = len(acumulado[canonico]["hrefs"])
        if ya >= target:
            cola.task_done()
            continue
        faltan = (target - ya) + 20
        res = await con_reintentos(
            lambda: recolectar_en_zona(page, termino, zona, faltan, zoom),
            reintentos, f"{termino}/{zona['nombre']}")
        if res == "BLOQUEO":
            if not await manejar_bloqueo(page, cfg):
                cola.task_done()
                continue
            res = await con_reintentos(
                lambda: recolectar_en_zona(page, termino, zona, faltan, zoom),
                reintentos, f"{termino}/{zona['nombre']}")
        hrefs = res if isinstance(res, list) else []
        async with lock:
            slot = acumulado[canonico]
            for h in hrefs:
                k = clave_lugar(h)
                if k not in slot["claves"] and len(slot["hrefs"]) < target:
                    slot["claves"].add(k)
                    slot["hrefs"].append((h, zona["nombre"]))
            log.info(f"  [{canonico}] '{termino}' @ {zona['nombre']}: "
                     f"{len(slot['hrefs'])} acumulados")
        cola.task_done()


# ============================================================
#  FASE 2 — extraer detalles + contactos + score
# ============================================================
async def extraer_negocio(page, href, canonico, zona):
    await page.goto(href, wait_until="domcontentloaded")
    await aceptar_consentimiento(page)
    try:
        await page.wait_for_selector("h1.DUwDvf, h1.fontHeadlineLarge", timeout=10000)
    except Exception:
        return None
    await page.wait_for_timeout(450)

    nombre = await _texto(page, ["h1.DUwDvf", "h1.fontHeadlineLarge", "h1"])
    bloque = await _texto(page, ["div.F7nice"])
    calificacion, resenas = "", ""
    m = re.search(r"([\d,\.]+)", bloque)
    if m:
        calificacion = m.group(1).replace(",", ".")
    m2 = re.search(r"\(([\d\.\,]+)\)", bloque)
    if m2:
        resenas = m2.group(1).replace(".", "").replace(",", "")
    categoria = await _texto(page, ["button.DkEaL", "button[jsaction*='category']"])
    direccion = limpiar_aria(
        await _attr(page, ['button[data-item-id="address"]'], "aria-label"),
        ["Dirección", "Address"])
    telefono = normalizar_tel(limpiar_aria(
        await _attr(page, ['button[data-item-id^="phone"]'], "aria-label"),
        ["Teléfono", "Phone"]))
    sitio_web = await _attr(page, ['a[data-item-id="authority"]'], "href")
    horario = limpiar_aria(
        await _attr(page, ['button[data-item-id="oh"]'], "aria-label"),
        ["Horario", "Hours"])
    coordenadas = ""
    mc = re.search(r"@(-?\d+\.\d+),(-?\d+\.\d+)", page.url)
    if mc:
        coordenadas = f"{mc.group(1)},{mc.group(2)}"

    datos = {"nombre": nombre, "categoria": categoria or canonico,
             "calificacion": calificacion, "resenas": resenas,
             "direccion": direccion, "telefono": telefono, "whatsapp": "",
             "email": "", "instagram": "", "facebook": "", "sitio_web": sitio_web,
             "sin_web": "", "score": 0, "horario": horario,
             "coordenadas": coordenadas, "zona": zona, "url_maps": href}

    # Si la "web" es en realidad una red social, reubícala
    low = sitio_web.lower()
    if "instagram.com/" in low:
        datos["instagram"] = sitio_web; datos["sitio_web"] = ""
    elif "facebook.com/" in low:
        datos["facebook"] = sitio_web; datos["sitio_web"] = ""
    elif "wa.me/" in low or "whatsapp.com" in low:
        mm = RE_WA.search(sitio_web)
        if mm:
            datos["whatsapp"] = normalizar_tel(mm.group(1))
        datos["sitio_web"] = ""
    return datos


async def worker_fase2(page, cola, escritor, lock, cliente, cfg):
    reintentos = cfg["rendimiento"].get("reintentos", 3)
    descanso = cfg["rendimiento"].get("descanso_cada", 40)
    quiere_email = cfg["extraccion"].get("extraer_email", True)
    quiere_redes = cfg["extraccion"].get("extraer_redes", True)
    timeout_web = cfg["extraccion"].get("timeout_web_seg", 10)
    while True:
        if PARAR.is_set():
            return
        try:
            canonico, href, zona = cola.get_nowait()
        except asyncio.QueueEmpty:
            return
        if escritor.ya_existe(href):
            cola.task_done()
            continue
        datos = await con_reintentos(
            lambda: extraer_negocio(page, href, canonico, zona),
            reintentos, "detalle")
        if datos is None and await detectar_bloqueo(page):
            if await manejar_bloqueo(page, cfg):
                datos = await con_reintentos(
                    lambda: extraer_negocio(page, href, canonico, zona),
                    reintentos, "detalle")
        if datos and datos["nombre"]:
            if (quiere_email or quiere_redes) and datos["sitio_web"] and cliente:
                cont = await extraer_contactos(cliente, datos["sitio_web"],
                                               timeout_web, quiere_redes)
                for k, v in cont.items():
                    if v and not datos.get(k):
                        datos[k] = v
            datos["sin_web"] = "" if datos["sitio_web"] else "SÍ"
            # verificación de contactos + actividad
            verificacion.enriquecer(datos)
            # web activa (no está caída)
            if datos["sitio_web"] and cliente:
                activa = await verificacion.web_responde(cliente, datos["sitio_web"])
                datos["web_activa"] = "sí" if activa else ("no" if activa is False else "")
            datos["score"] = calcular_score(datos)
            async with lock:
                escritor.escribir(canonico, datos)
                n = escritor.conteo[canonico]
                total = escritor.total
                PROGRESO["total"] = total
            if ON_LEAD:
                try:
                    ON_LEAD(dict(datos, _nicho=canonico))
                except Exception:
                    pass
            marca = ("🟢" if datos["score"] >= 5 else
                     "🟡" if datos["score"] >= 3 else "⚪")
            log.info(f"  [{canonico}] {n}. {marca} {datos['nombre']} "
                     f"(score {datos['score']})")
            if descanso and total % descanso == 0:
                await page.wait_for_timeout(random.randint(4000, 8000))
        base = 600 + random.randint(0, 600)
        await page.wait_for_timeout(base * 3 if LENTO else base)
        cola.task_done()


# ============================================================
#  FASE 3 — Excel (Dashboard + hojas ordenadas por score)
# ============================================================
def _leer_csv(ruta):
    with open(ruta, encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def construir_excel(carpeta_temp, archivo_salida):
    wb = Workbook(write_only=True)
    fill_h = PatternFill("solid", fgColor="1F4E78")
    fill_d = PatternFill("solid", fgColor="2E7D32")
    font_h = Font(bold=True, color="FFFFFF")
    font_link = Font(color="1155CC", underline="single")

    archivos = sorted(f for f in os.listdir(carpeta_temp) if f.endswith(".csv"))
    stats = {}

    # ---- Dashboard ----
    ws = wb.create_sheet(title="Dashboard")
    for i, w in enumerate([24, 10, 12, 12, 12, 12, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    enc = ["Nicho", "Total", "Con tel", "WhatsApp", "Email", "Redes", "Sin web"]
    fila = []
    for h in enc:
        c = WriteOnlyCell(ws, value=h); c.fill = fill_d; c.font = font_h
        fila.append(c)
    ws.append(fila)
    tot = dict(total=0, tel=0, wa=0, mail=0, red=0, sinweb=0)
    datos_por_nicho = {}
    for archivo in archivos:
        filas = _leer_csv(os.path.join(carpeta_temp, archivo))
        for f in filas:
            try:
                f["_score"] = float(f.get("score") or 0)
            except Exception:
                f["_score"] = 0
        filas.sort(key=lambda x: x["_score"], reverse=True)
        datos_por_nicho[archivo] = filas
        s = dict(total=len(filas),
                 tel=sum(1 for f in filas if f.get("telefono")),
                 wa=sum(1 for f in filas if f.get("whatsapp")),
                 mail=sum(1 for f in filas if f.get("email")),
                 red=sum(1 for f in filas if f.get("instagram") or f.get("facebook")),
                 sinweb=sum(1 for f in filas if f.get("sin_web")))
        stats[archivo[:-4]] = s
        ws.append([archivo[:-4], s["total"], s["tel"], s["wa"],
                   s["mail"], s["red"], s["sinweb"]])
        for k in tot:
            tot[k] += s[k]
    frow = [WriteOnlyCell(ws, value=v) for v in
            ["TOTAL", tot["total"], tot["tel"], tot["wa"],
             tot["mail"], tot["red"], tot["sinweb"]]]
    for c in frow:
        c.font = Font(bold=True)
    ws.append(frow)

    # ---- Hojas por nicho (ordenadas por score) ----
    idx = {n: i for i, n in enumerate(CAMPOS)}
    for archivo in archivos:
        nicho = archivo[:-4]
        ws = wb.create_sheet(title=nicho[:31])
        anchos = [28, 18, 8, 16, 16, 10, 26, 11, 22, 22, 30, 10, 8, 12,
                  11, 9, 34, 22, 20, 14, 40]
        for i, w in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        fila = []
        for h in ENCABEZADOS:
            c = WriteOnlyCell(ws, value=h)
            c.fill = fill_h; c.font = font_h
            c.alignment = Alignment(vertical="center")
            fila.append(c)
        ws.append(fila)
        for f in datos_por_nicho[archivo]:
            celdas = []
            for campo in CAMPOS:
                val = f.get(campo, "")
                c = WriteOnlyCell(ws, value=val)
                if val and campo == "telefono":
                    c.hyperlink = "tel:" + re.sub(r"[^\d+]", "", val); c.font = font_link
                elif val and campo == "whatsapp":
                    c.hyperlink = "https://wa.me/" + re.sub(r"[^\d]", "", val)
                    c.font = font_link
                elif val and campo == "email":
                    c.hyperlink = "mailto:" + val.split(";")[0].strip(); c.font = font_link
                elif val and campo in ("instagram", "facebook", "sitio_web"):
                    u = val if val.startswith("http") else "https://" + val
                    c.hyperlink = u; c.font = font_link
                celdas.append(c)
            ws.append(celdas)

    if not archivos:
        wb.create_sheet(title="Sin datos")
    wb.save(archivo_salida)
    return stats


# ============================================================
#  ORQUESTACIÓN
# ============================================================
async def run(cfg):
    global LENTO
    t_inicio = time.time()
    b, r, s = cfg["busqueda"], cfg["rendimiento"], cfg["salida"]
    nichos = parse_nichos(b["nichos"])
    zonas = resolver_zonas(b)

    # Modo prueba: corrida mínima para verificar que todo funciona en segundos.
    if r.get("modo_prueba"):
        nichos = nichos[:1]
        zonas = zonas[:1]
        b["target_por_nicho"] = min(int(b.get("target_por_nicho", 100) or 100), 8)
        r["concurrencia"] = 1
        log.info("🧪 MODO PRUEBA: 1 nicho, 1 zona, hasta "
                 f"{b['target_por_nicho']} resultados.")

    # Modo prudente: fuerza 1 pestaña y ritmo lento (menos captchas)
    LENTO = bool(r.get("modo_prudente", False))
    concurrencia = 1 if LENTO else r.get("concurrencia", 2)

    escritor = EscritorCSV(s["carpeta_temp"])
    acumulado = {can: {"hrefs": [], "claves": set()} for can, _ in nichos}
    lock = asyncio.Lock()
    hacer_web = cfg["extraccion"].get("extraer_email", True) or \
        cfg["extraccion"].get("extraer_redes", True)
    cliente = httpx.AsyncClient(headers={"User-Agent": USER_AGENT},
                                verify=False) if hacer_web else None

    args = ["--disable-blink-features=AutomationControlled"]
    user_data_dir = r.get("user_data_dir")   # sesión persistente (cookies)

    try:
        async with async_playwright() as p:
            navegador = None
            contexto = None
            try:
                if user_data_dir:
                    os.makedirs(user_data_dir, exist_ok=True)
                    contexto = await p.chromium.launch_persistent_context(
                        user_data_dir, headless=r.get("headless", False),
                        locale="es-ES", user_agent=USER_AGENT, args=args)
                else:
                    navegador = await p.chromium.launch(
                        headless=r.get("headless", False), args=args)
                    contexto = await navegador.new_context(locale="es-ES",
                                                           user_agent=USER_AGENT)
                await contexto.add_init_script(STEALTH_JS)
                # Cookie que evita la pantalla de consentimiento de Google.
                try:
                    await contexto.add_cookies([{
                        "name": "CONSENT",
                        "value": f"YES+cb.20210328-17-p0.es+FX+{random.randint(100, 999)}",
                        "domain": ".google.com",
                        "path": "/",
                    }])
                except Exception:
                    pass
                existentes = contexto.pages
                paginas = list(existentes[:concurrencia])
                while len(paginas) < concurrencia:
                    paginas.append(await contexto.new_page())

                if LENTO:
                    log.info("🐢 Modo prudente activo: 1 pestaña, ritmo lento.")
                log.info(f"=== FASE 1: {len(nichos)} nichos × {len(zonas)} zonas ===")
                cola1 = asyncio.Queue()
                for canonico, terminos in nichos:
                    for termino in terminos:
                        for zona in zonas:
                            cola1.put_nowait((canonico, termino, zona))
                try:
                    await asyncio.gather(*[worker_fase1(pg, cola1, acumulado, lock, cfg)
                                           for pg in paginas], return_exceptions=True)
                except Exception as e:
                    log.warning(f"Aviso en fase 1: {e}")

                if not PARAR.is_set():
                    log.info("=== FASE 2: extrayendo detalles + contactos ===")
                    cola2 = asyncio.Queue()
                    for canonico, slot in acumulado.items():
                        for href, zona in slot["hrefs"]:
                            cola2.put_nowait((canonico, href, zona))
                    try:
                        await asyncio.gather(*[worker_fase2(pg, cola2, escritor, lock, cliente, cfg)
                                               for pg in paginas], return_exceptions=True)
                    except Exception as e:
                        log.warning(f"Aviso en fase 2: {e}")
            except Exception as e:
                log.error(f"Ocurrió un problema durante la búsqueda: {e}")
            finally:
                try:
                    if contexto:
                        await contexto.close()
                except Exception:
                    pass
                try:
                    if navegador:
                        await navegador.close()
                except Exception:
                    pass
    except Exception as e:
        log.error(f"No se pudo iniciar el navegador: {e}")
    finally:
        if cliente:
            try:
                await cliente.aclose()
            except Exception:
                pass
        try:
            escritor.cerrar()
        except Exception:
            pass

    log.info("=== Generando Excel ===")
    try:
        stats = construir_excel(s["carpeta_temp"], s["archivo_excel"])
    except Exception as e:
        log.error(f"No se pudo generar el Excel: {e}")
        stats = {}
    total = sum(v["total"] for v in stats.values())
    dur = time.time() - t_inicio
    log.info("=" * 50)
    if PARAR.is_set():
        log.info(f"⏹ Detenido — {total} negocios guardados hasta ahora.")
    else:
        log.info(f"✅ LISTO — {total} negocios en '{s['archivo_excel']}'")
    log.info(f"⏱ Tiempo total: {dur:.1f}s")
    for nicho, v in stats.items():
        log.info(f"   {nicho}: {v['total']} (tel {v['tel']}, wa {v['wa']}, "
                 f"email {v['mail']}, redes {v['red']}, sin web {v['sinweb']})")
    log.info("=" * 50)


def main():
    configurar_logging()
    ruta = sys.argv[1] if len(sys.argv) > 1 else "config.yaml"
    asyncio.run(run(cargar_config(ruta)))


if __name__ == "__main__":
    main()
